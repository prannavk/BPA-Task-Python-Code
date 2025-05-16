from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from .remote_data_fetch import fetch_and_parse_suppliers

def write_suppliers_to_excel(suppliers: List[Dict[str, str]], save_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    if not suppliers:
        print("Empty supplier list. Nothing to write.")
        return

    keys = list(suppliers[0].keys())  # Header from keys
    num_cols = len(keys)
    last_col_letter = get_column_letter(num_cols)

    # Title Row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Northwind Suppliers"
    title_cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Header Row
    header_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for col, key in enumerate(keys, start=1):
        cell = ws.cell(row=2, column=col, value=key)
        cell.fill = header_fill
        cell.font = Font(bold=True)

    # Data Rows
    for row_idx, supplier in enumerate(suppliers, start=3):
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=supplier.get(key, ""))

    # Save the workbook
    wb.save(save_path)
    print(f"Excel file saved at: {save_path}")


if __name__ == "__main__":
    print("ok steps 1, 2\n")    
    write_suppliers_to_excel(fetch_and_parse_suppliers(), "Northwind_Suppliers.xlsx")
    
